import pandas as pd
import numpy as np
from openpyxl import load_workbook
import json

def clean_value(val):
    """Clean and convert values, similar to excel_to_js.py"""
    if val is None or (isinstance(val, str) and val.strip() == ''):
        return None
    
    # If it's already a number, preserve its type and precision
    if isinstance(val, (int, float)):
        return val
    
    # If it's a string, clean and convert
    if isinstance(val, str):
        val = val.replace('"', '').replace(',', '').replace("'", '').replace('"', '').replace('"', '')
        try:
            # Try float first to preserve decimals
            float_val = float(val)
            # Only convert to int if it's a whole number
            if float_val.is_integer():
                return int(float_val)
            else:
                return float_val
        except (ValueError, TypeError):
            return val
    
    return val

def clean_excel_data(file_path):
    """
    Clean up the messy Turkish electricity data Excel file
    Handle red-colored values by wrapping them in parentheses
    """
    print(f"🧹 CLEANING EXCEL DATA: {file_path}")
    print("=" * 60)
    
    # Load workbook with openpyxl to access formatting
    # First load with data_only=True to get calculated values
    wb_data = load_workbook(file_path, data_only=True)
    # Then load with data_only=False to access formatting
    wb_format = load_workbook(file_path, data_only=False)
    
    cleaned_data = {}
    
    for sheet_name in wb_data.sheetnames:
        ws_data = wb_data[sheet_name]
        ws_format = wb_format[sheet_name]
        print(f"\n📊 Cleaning sheet: '{sheet_name}'")
        print("-" * 40)
        
        # Get all rows with values and formatting
        rows_data = list(ws_data.iter_rows(values_only=False))
        rows_format = list(ws_format.iter_rows(values_only=False))
        if not rows_data or not rows_format:
            continue
            
        # Extract header row
        header_row = rows_data[0]
        headers = []
        for i, cell in enumerate(header_row):
            if cell.value is None or str(cell.value).strip() == '':
                headers.append(f'Column_{i+1}')
            else:
                clean_name = str(cell.value).strip()
                clean_name = clean_name.replace('   ', ' ').replace('  ', ' ')
                headers.append(clean_name)
        
        # Process data rows
        data_rows = []
        for i, (row_data, row_format) in enumerate(zip(rows_data[1:], rows_format[1:]), 1):
            if not row_data or not row_data[0].value or str(row_data[0].value).strip() == '':
                continue
                
            row_values = []
            for j, (cell_data, cell_format) in enumerate(zip(row_data, row_format)):
                if j >= len(headers):
                    break
                    
                # Get calculated value from data workbook
                val = cell_data.value
                cleaned = clean_value(val)
                
                # Check for red font using format workbook
                is_red = False
                if cell_format and cell_format.font and cell_format.font.color:
                    rgb = getattr(cell_format.font.color, 'rgb', None)
                    if rgb:
                        rgb_str = str(rgb)
                        if rgb_str.upper().endswith('FF0000') or rgb_str.upper() == 'FFFF0000':
                            is_red = True
                
                # Wrap red values in parentheses
                if is_red and cleaned is not None:
                    if isinstance(cleaned, (int, float)):
                        row_values.append(f'({cleaned})')
                    else:
                        row_values.append(f'({cleaned})')
                else:
                    row_values.append(cleaned)
            
            if row_values:  # Only add non-empty rows
                data_rows.append(row_values)
        
        # Create DataFrame
        if data_rows:
            # Ensure all rows have the same length
            max_cols = len(headers)
            for i, row in enumerate(data_rows):
                while len(row) < max_cols:
                    row.append(None)
                data_rows[i] = row[:max_cols]
            
            df = pd.DataFrame(data_rows, columns=headers)
            
            # Store original dimensions
            orig_rows, orig_cols = df.shape
            
            # Clean the year column (first column) - handle parentheses values
            print("   📅 Cleaning year column...")
            year_col = df.columns[0]
            
            # Extract numeric values from parentheses if needed for year column
            year_values = []
            for val in df[year_col]:
                if isinstance(val, str) and val.startswith('(') and val.endswith(')'):
                    # Extract number from parentheses for year
                    try:
                        year_values.append(int(val.strip('()')))
                    except:
                        year_values.append(None)
                else:
                    try:
                        year_values.append(int(val) if val is not None else None)
                    except:
                        year_values.append(None)
            
            df[year_col] = year_values
            
            # Remove rows with missing years
            df = df.dropna(subset=[year_col])
            
            # Ensure years are integers
            df[year_col] = df[year_col].astype('Int64')
            
            # For other columns, convert non-parentheses values to numeric
            print("   🔢 Cleaning numeric data...")
            for col in df.columns[1:]:
                numeric_values = []
                for val in df[col]:
                    if isinstance(val, str) and val.startswith('(') and val.endswith(')'):
                        # Keep parentheses values as strings
                        numeric_values.append(val)
                    else:
                        # Try to convert to numeric
                        try:
                            if val is not None and str(val).strip() != '':
                                numeric_val = float(val)
                                # Round to 3 decimal places
                                numeric_values.append(round(numeric_val, 3))
                            else:
                                numeric_values.append(None)
                        except:
                            numeric_values.append(val)
                
                df[col] = numeric_values
            
            # Sort by year
            df = df.sort_values(by=year_col).reset_index(drop=True)
            
            # Remove duplicate years if any
            df = df.drop_duplicates(subset=[year_col], keep='first')
            
            cleaned_data[sheet_name] = df
            
            # Report cleaning results
            new_rows, new_cols = df.shape
            print(f"   ✅ Results:")
            print(f"      • Rows: {orig_rows} → {new_rows} ({new_rows-orig_rows:+d})")
            print(f"      • Columns: {orig_cols} → {new_cols} ({new_cols-orig_cols:+d})")
            print(f"      • Year range: {df[year_col].min()} - {df[year_col].max()}")
            
            # Count parentheses values
            parentheses_count = 0
            for col in df.columns[1:]:
                parentheses_count += df[col].astype(str).str.contains(r'^\(.*\)$', na=False).sum()
            print(f"      • Red values (in parentheses): {parentheses_count}")
    
    wb_data.close()
    wb_format.close()
    return cleaned_data

def save_cleaned_data(cleaned_data, output_prefix="cleaned"):
    """
    Save cleaned data in multiple formats
    """
    print(f"\n💾 SAVING CLEANED DATA")
    print("=" * 60)
    
    # 1. Save as Excel
    excel_file = f"{output_prefix}_electricity_data.xlsx"
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        for sheet_name, df in cleaned_data.items():
            # Create a clean sheet name for Excel
            clean_sheet_name = sheet_name.replace('/', '_').replace('\\', '_')[:31]  # Excel limit
            df.to_excel(writer, sheet_name=clean_sheet_name, index=False)
    
    print(f"   ✅ Excel file saved: {excel_file}")
    
    # 2. Save as CSV files
    for sheet_name, df in cleaned_data.items():
        csv_file = f"{output_prefix}_{sheet_name.lower().replace(' ', '_')}.csv"
        df.to_csv(csv_file, index=False, encoding='utf-8')
        print(f"   ✅ CSV file saved: {csv_file}")
    
    # 3. Save as JSON (for web use)
    json_data = {}
    for sheet_name, df in cleaned_data.items():
        # Convert to records format
        json_data[sheet_name] = df.to_dict('records')
    
    json_file = f"{output_prefix}_electricity_data.json"
    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump(json_data, f, ensure_ascii=False, indent=2)
    
    print(f"   ✅ JSON file saved: {json_file}")
    
    return excel_file, json_file

def generate_summary_report(cleaned_data):
    """
    Generate a summary report of the cleaned data
    """
    print(f"\n📋 DATA SUMMARY REPORT")
    print("=" * 60)
    
    for sheet_name, df in cleaned_data.items():
        print(f"\n📊 Sheet: '{sheet_name}'")
        print("-" * 30)
        
        # Basic info
        print(f"   📏 Dimensions: {df.shape[0]} rows × {df.shape[1]} columns")
        print(f"   📅 Time period: {df.iloc[0, 0]} - {df.iloc[-1, 0]} ({df.iloc[-1, 0] - df.iloc[0, 0] + 1} years)")
        
        # Column summary
        print(f"   📊 Columns:")
        for i, col in enumerate(df.columns):
            if i == 0:
                print(f"      {i+1:2d}. {col} (Year column)")
            else:
                # Handle mixed data types (numeric and parentheses strings)
                non_zero = 0
                parentheses_count = 0
                total_vals = df[col].notna().sum()
                
                for val in df[col]:
                    if pd.notna(val):
                        if isinstance(val, str) and val.startswith('(') and val.endswith(')'):
                            parentheses_count += 1
                            # Extract numeric value from parentheses to check if > 0
                            try:
                                num_val = float(val.strip('()'))
                                if num_val > 0:
                                    non_zero += 1
                            except:
                                pass
                        elif isinstance(val, (int, float)) and val > 0:
                            non_zero += 1
                
                if parentheses_count > 0:
                    print(f"      {i+1:2d}. {col} ({non_zero}/{total_vals} non-zero values, {parentheses_count} in parentheses)")
                else:
                    print(f"      {i+1:2d}. {col} ({non_zero}/{total_vals} non-zero values)")
        
        # Data quality metrics
        missing_data = df.isna().sum().sum()
        total_cells = df.shape[0] * df.shape[1]
        completeness = ((total_cells - missing_data) / total_cells) * 100
        
        print(f"   📈 Data Quality:")
        print(f"      • Completeness: {completeness:.1f}% ({total_cells - missing_data}/{total_cells} cells)")
        print(f"      • Missing values: {missing_data}")
        
        # Sample recent data
        print(f"   🔍 Recent data (last 3 years):")
        for idx in range(max(0, len(df)-3), len(df)):
            year = df.iloc[idx, 0]
            total_col = None
            for col in df.columns:
                if 'toplam' in str(col).lower() and 'termik' not in str(col).lower():
                    total_col = col
                    break
            
            if total_col:
                total_val = df.iloc[idx][total_col]
                print(f"      {year}: Total = {total_val}")

def main():
    # Input file
    input_file = "b elektrik generjisinin kaynaklara göre kurulu gücü ve üretimi.xlsx"
    
    print("🚀 TURKISH ELECTRICITY DATA CLEANUP TOOL")
    print("=" * 80)
    print(f"Input file: {input_file}")
    print()
    
    # Step 1: Clean the data
    cleaned_data = clean_excel_data(input_file)
    
    # Step 2: Save in multiple formats
    excel_file, json_file = save_cleaned_data(cleaned_data)
    
    # Step 3: Generate summary report
    generate_summary_report(cleaned_data)
    
    print(f"\n{'='*80}")
    print("🎉 CLEANUP COMPLETE!")
    print(f"   📁 Clean Excel file: {excel_file}")
    print(f"   📁 JSON file: {json_file}")
    print(f"   📁 CSV files: cleaned_*.csv")
    print()
    print("💡 The data is now:")
    print("   • Properly formatted with consistent precision")
    print("   • Free of empty columns and problematic headers")
    print("   • Sorted chronologically")
    print("   • Available in multiple formats (Excel, CSV, JSON)")
    print("   • Red values wrapped in parentheses")

if __name__ == "__main__":
    main() 