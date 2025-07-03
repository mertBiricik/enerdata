#!/usr/bin/env python3
"""
Excel File Structure Analyzer
Analyzes the structure of Excel files to understand their layout, data patterns, and potential issues.
"""

import pandas as pd
import numpy as np
import openpyxl
from openpyxl import load_workbook
import sys
import os
from collections import Counter
import warnings
warnings.filterwarnings('ignore')

def analyze_excel_structure(file_path):
    """
    Comprehensive analysis of Excel file structure
    """
    print(f"🔍 Analyzing Excel file: {file_path}")
    print("=" * 60)
    
    if not os.path.exists(file_path):
        print(f"❌ File not found: {file_path}")
        return
    
    # 1. Basic file info
    file_size = os.path.getsize(file_path)
    print(f"📁 File size: {file_size / 1024:.2f} KB")
    
    try:
        # Load with openpyxl for detailed analysis
        wb = load_workbook(file_path, data_only=False)
        print(f"📊 Number of worksheets: {len(wb.worksheets)}")
        
        for i, ws in enumerate(wb.worksheets):
            print(f"\n📋 Sheet {i+1}: '{ws.title}'")
            print("-" * 40)
            
            # Basic dimensions
            max_row = ws.max_row
            max_col = ws.max_column
            print(f"   📏 Dimensions: {max_row} rows × {max_col} columns")
            
            # Get actual data range (non-empty cells)
            used_range = get_used_range(ws)
            print(f"   📊 Used range: {used_range}")
            
            # Analyze first few rows to understand structure
            print(f"\n   🔍 First 10 rows analysis:")
            analyze_rows(ws, max_rows=10)
            
            # Check for merged cells
            merged_ranges = list(ws.merged_cells.ranges)
            if merged_ranges:
                print(f"   🔗 Merged cell ranges: {len(merged_ranges)}")
                for merge in merged_ranges[:5]:  # Show first 5
                    print(f"      - {merge}")
                if len(merged_ranges) > 5:
                    print(f"      ... and {len(merged_ranges) - 5} more")
            
            # Check for formulas
            formula_cells = find_formula_cells(ws)
            if formula_cells:
                print(f"   🧮 Formula cells: {len(formula_cells)}")
                for cell, formula in formula_cells[:3]:
                    print(f"      - {cell}: {formula}")
                if len(formula_cells) > 3:
                    print(f"      ... and {len(formula_cells) - 3} more")
            
            # Data type analysis
            print(f"\n   📈 Data type analysis:")
            analyze_data_types(ws)
            
            # Look for patterns and issues
            print(f"\n   ⚠️  Potential issues:")
            identify_issues(ws)
            
        # Try pandas analysis for data insights
        print(f"\n🐼 Pandas-based analysis:")
        analyze_with_pandas(file_path)
        
    except Exception as e:
        print(f"❌ Error analyzing file: {str(e)}")
        print(f"   Trying basic pandas read...")
        try:
            df = pd.read_excel(file_path, header=None)
            print(f"   Shape: {df.shape}")
            print(f"   First 5 rows:")
            print(df.head())
        except Exception as e2:
            print(f"   ❌ Pandas also failed: {str(e2)}")

def get_used_range(ws):
    """Get the actual used range of the worksheet"""
    min_row, max_row = 1, 1
    min_col, max_col = 1, 1
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                min_row = min(min_row, cell.row) if min_row > 1 else cell.row
                max_row = max(max_row, cell.row)
                min_col = min(min_col, cell.column) if min_col > 1 else cell.column
                max_col = max(max_col, cell.column)
    
    return f"{openpyxl.utils.get_column_letter(min_col)}{min_row}:{openpyxl.utils.get_column_letter(max_col)}{max_row}"

def analyze_rows(ws, max_rows=10):
    """Analyze the first few rows to understand structure"""
    for row_num in range(1, min(max_rows + 1, ws.max_row + 1)):
        row_data = []
        for col_num in range(1, min(11, ws.max_column + 1)):  # First 10 columns
            cell = ws.cell(row_num, col_num)
            value = cell.value
            if value is not None:
                # Truncate long values
                str_val = str(value)
                if len(str_val) > 30:
                    str_val = str_val[:27] + "..."
                row_data.append(str_val)
            else:
                row_data.append("[empty]")
        
        print(f"      Row {row_num}: {' | '.join(row_data[:5])}")
        if len(row_data) > 5:
            print(f"              ... and {len(row_data) - 5} more columns")

def find_formula_cells(ws):
    """Find cells containing formulas"""
    formulas = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == 'f':  # Formula
                formulas.append((cell.coordinate, cell.value))
    return formulas

def analyze_data_types(ws):
    """Analyze data types in the worksheet"""
    type_counts = Counter()
    total_cells = 0
    empty_cells = 0
    
    for row in ws.iter_rows():
        for cell in row:
            total_cells += 1
            if cell.value is None:
                empty_cells += 1
                type_counts['empty'] += 1
            elif isinstance(cell.value, str):
                type_counts['text'] += 1
            elif isinstance(cell.value, (int, float)):
                type_counts['number'] += 1
            elif hasattr(cell.value, 'date'):  # datetime
                type_counts['date'] += 1
            else:
                type_counts['other'] += 1
    
    print(f"      Total cells: {total_cells}")
    print(f"      Empty cells: {empty_cells} ({empty_cells/total_cells*100:.1f}%)")
    for data_type, count in type_counts.most_common():
        if data_type != 'empty':
            print(f"      {data_type.capitalize()}: {count} ({count/total_cells*100:.1f}%)")

def identify_issues(ws):
    """Identify potential structural issues"""
    issues = []
    
    # Check for inconsistent row lengths
    row_lengths = []
    for row_num in range(1, min(21, ws.max_row + 1)):  # Check first 20 rows
        row_length = 0
        for col_num in range(1, ws.max_column + 1):
            if ws.cell(row_num, col_num).value is not None:
                row_length = col_num
        row_lengths.append(row_length)
    
    if len(set(row_lengths)) > 3:  # More than 3 different row lengths
        issues.append(f"Inconsistent row lengths: {set(row_lengths)}")
    
    # Check for mixed data types in columns
    for col_num in range(1, min(11, ws.max_column + 1)):  # First 10 columns
        col_types = set()
        for row_num in range(1, min(21, ws.max_row + 1)):  # First 20 rows
            cell_value = ws.cell(row_num, col_num).value
            if cell_value is not None:
                col_types.add(type(cell_value).__name__)
        
        if len(col_types) > 2:  # More than 2 types (allowing for some flexibility)
            col_letter = openpyxl.utils.get_column_letter(col_num)
            issues.append(f"Column {col_letter} has mixed data types: {col_types}")
    
    if issues:
        for issue in issues:
            print(f"      - {issue}")
    else:
        print(f"      - No obvious structural issues detected")

def analyze_with_pandas(file_path):
    """Use pandas for additional analysis"""
    try:
        # Try different reading strategies
        strategies = [
            {"header": 0},
            {"header": None},
            {"header": 1},
            {"header": [0, 1]},  # Multi-level header
        ]
        
        for i, strategy in enumerate(strategies):
            try:
                df = pd.read_excel(file_path, **strategy)
                print(f"   Strategy {i+1} ({strategy}): Success!")
                print(f"      Shape: {df.shape}")
                print(f"      Columns: {list(df.columns)[:5]}...")
                
                # Check for obvious patterns
                if df.shape[1] > 0:
                    first_col = df.iloc[:, 0]
                    print(f"      First column sample: {first_col.head(3).tolist()}")
                
                break
            except Exception as e:
                print(f"   Strategy {i+1} ({strategy}): Failed - {str(e)[:50]}...")
        
    except Exception as e:
        print(f"   ❌ Pandas analysis failed: {str(e)}")

def main():
    if len(sys.argv) != 2:
        print("Usage: python analyze_excel_structure.py <excel_file_path>")
        sys.exit(1)
    
    file_path = sys.argv[1]
    analyze_excel_structure(file_path)

if __name__ == "__main__":
    main() 