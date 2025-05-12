import json
from openpyxl import load_workbook
from openpyxl.styles.colors import COLOR_INDEX

def clean_value(val):
    if val is None or (isinstance(val, str) and val.strip() == ''):
        return None
    if isinstance(val, str):
        val = val.replace('"', '').replace(',', '').replace('’', '').replace('"', '').replace('"', '')
    try:
        if isinstance(val, str) and '.' in val:
            return float(val)
        return int(val)
    except (ValueError, TypeError):
        return val

input_xlsx = 'source.xlsx'
output_js = 'embedded_data.js'

wb = load_workbook(input_xlsx, data_only=True)
ws = wb.active  # First sheet

rows = list(ws.iter_rows(values_only=False))
header = [cell.value for cell in rows[0]]
years = [int(y) for y in header[1:] if y and str(y).strip() != '']

embedded_data = []
for row in rows[1:]:
    if not row or not row[0].value or str(row[0].value).strip() == '':
        continue
    category = str(row[0].value).strip().replace('"', '').replace('"', '').replace('"', '')
    if not category:
        continue
    obj = {'category': category}
    for i, year in enumerate(years):
        cell = row[i+1] if i+1 < len(row) else None
        val = cell.value if cell else None
        cleaned = clean_value(val)
        # Check for red font (font color is usually 'FFFF0000' or 'FF0000' for red)
        is_red = False
        if cell and cell.font and cell.font.color:
            rgb = getattr(cell.font.color, 'rgb', None)
            if rgb:
                rgb_str = str(rgb)
                if rgb_str.upper().endswith('FF0000') or rgb_str.upper() == 'FFFF0000':
                    is_red = True
        if is_red and cleaned is not None:
            obj[year] = f'({cleaned})'
        else:
            obj[year] = cleaned
    embedded_data.append(obj)

with open(output_js, 'w', encoding='utf-8') as f:
    f.write('const embeddedRawData = ')
    json.dump(embedded_data, f, ensure_ascii=False, indent=2)
    f.write(';')

print(f"Done! JS array written to {output_js}. Copy its contents into your index.html.") 